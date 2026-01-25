"""
BOM管理API
"""
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, text
from typing import Optional
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from decimal import Decimal
import os
import re
from datetime import datetime
import json
from urllib.parse import quote
try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False
from app.core.database import get_db

# Debug logging setup
DEBUG_LOG_PATH = r"d:\dream-start\code\procurement-manager\.cursor\debug.log"

def _log_debug(location: str, message: str, data: dict, hypothesis_id: str = "A"):
    """Write debug log to file"""
    try:
        log_entry = {
            "sessionId": "debug-session",
            "runId": "run1",
            "hypothesisId": hypothesis_id,
            "location": location,
            "message": message,
            "data": data,
            "timestamp": int(datetime.now().timestamp() * 1000)
        }
        with open(DEBUG_LOG_PATH, "a", encoding="utf-8") as f:
            f.write(json.dumps(log_entry, ensure_ascii=False) + "\n")
    except Exception:
        pass  # Silently fail if logging fails
from app.models.bom import BOM, BOMItem
from app.models.material import Material
from app.models.supplier import Supplier
from app.models.quotation import Quotation, QuotationItem
from app.schemas.bom import (
    BOMCreate, BOMUpdate, BOM as BOMSchema, BOMListResponse, ExcelImportResponse,
    BOMItemPriceHistoryResponse, BOMItemPriceHistoryItem, MaterialMatchInfo, PriceStatistics,
    BOMCostAnalysisResponse, BOMItemCostInfo
)
from app.schemas.quotation import QuotationComparisonResponse, QuotationBasicInfo, ComparisonItemRow, ComparisonItemCell
from app.core.dependencies import get_current_user
from app.models.user import User
from sqlalchemy.orm import joinedload
from sqlalchemy import func

router = APIRouter(prefix="/boms", tags=["BOM管理"])


def match_material_by_name_and_spec(
    db: Session,
    material_name: str,
    specification: Optional[str] = None
) -> Optional[Material]:
    """
    根据物料名称和规格精确匹配物料
    使用精确匹配：Material.name == material_name AND Material.specification == specification
    返回第一个匹配的物料（用于向后兼容）
    """
    query = db.query(Material).filter(Material.name == material_name)
    
    if specification:
        query = query.filter(Material.specification == specification)
    else:
        # 如果规格为空，则匹配规格也为空的物料
        query = query.filter(
            (Material.specification == None) | (Material.specification == "")
        )
    
    # 如果匹配到多个，选择第一个
    return query.first()


def _remove_all_spaces(s: Optional[str]) -> str:
    """去除字符串中的所有空格"""
    if s is None:
        return ""
    return s.replace(" ", "").replace("\t", "").replace("\n", "").replace("\r", "")


def match_all_materials_by_name_and_spec(
    db: Session,
    material_name: str,
    specification: Optional[str] = None
) -> list[Material]:
    """
    根据物料名称和规格匹配所有物料
    返回所有匹配的物料记录
    注意：匹配时需要考虑物料名称和规格可能包含空格的情况
    数据库中存储的物料名称和规格已经去除了空格，但BOM明细项中的名称可能还包含空格
    """
    # 去除BOM明细项中的空格（因为数据库中存储的物料已去除空格）
    material_name_clean = _remove_all_spaces(material_name)
    spec_clean = _remove_all_spaces(specification) if specification else None
    
    # 查询所有物料，然后在Python中进行匹配（因为需要去除空格比较）
    # 如果数据量很大，可以考虑使用SQL的REPLACE函数，但为了兼容性和清晰度，使用Python过滤
    all_materials = db.query(Material).all()
    
    matched_materials = []
    for material in all_materials:
        # 去除物料名称中的空格进行比较
        material_name_db = _remove_all_spaces(material.name)
        
        if material_name_db == material_name_clean:
            # 名称匹配，检查规格
            if spec_clean:
                # 规格不为空，需要精确匹配
                spec_db = _remove_all_spaces(material.specification)
                if spec_db == spec_clean:
                    matched_materials.append(material)
            else:
                # 规格为空，匹配规格也为空的物料
                spec_db = _remove_all_spaces(material.specification)
                if not spec_db:
                    matched_materials.append(material)
    
    # 按更新时间降序排列（最新的在前）
    matched_materials.sort(key=lambda m: (m.updated_at or m.created_at or datetime.min), reverse=True)
    
    return matched_materials


def generate_bom_code(db: Session) -> str:
    """
    自动生成BOM编码
    格式：B + 日期(YYYYMMDD) + 序号(001, 002...)
    例如：B20250101-001
    """
    today = datetime.now().strftime("%Y%m%d")
    prefix = f"B{today}"
    
    # 查询今天已生成的BOM（编码格式为 B日期序号）
    # 使用前缀匹配
    today_boms = db.query(BOM).filter(
        BOM.code.like(f"{prefix}%")
    ).all()
    
    # 提取已有的序号
    existing_sequences = []
    for b in today_boms:
        try:
            # 编码格式：B20250101-001，提取最后3位
            if len(b.code) >= len(prefix) + 3 and b.code.startswith(prefix):
                seq_str = b.code[len(prefix):]
                # 确保后续部分是数字
                if seq_str.isdigit():
                    seq = int(seq_str)
                    existing_sequences.append(seq)
        except (ValueError, IndexError):
            continue
    
    # 生成序号（从001开始，找到第一个未使用的序号）
    sequence = 1
    while sequence in existing_sequences:
        sequence += 1
    
    code = f"{prefix}-{sequence:03d}"
    
    # 双重检查：确保编码唯一
    while db.query(BOM).filter(BOM.code == code).first():
        sequence += 1
        code = f"{prefix}-{sequence:03d}"
    
    return code


@router.get("", response_model=BOMListResponse)
async def get_boms(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    keyword: Optional[str] = None,
    product_name: Optional[str] = None,
    status: Optional[str] = None,
    customer_name: Optional[str] = None,
    prepared_by: Optional[str] = None,
    created_at_start: Optional[datetime] = None,
    created_at_end: Optional[datetime] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取BOM列表"""
    query = db.query(BOM).options(joinedload(BOM.items))
    
    # 搜索
    if keyword:
        query = query.filter(
            or_(
                BOM.code.like(f"%{keyword}%"),
                BOM.name.like(f"%{keyword}%")
            )
        )
    
    if product_name:
        query = query.filter(BOM.product_name.like(f"%{product_name}%"))
    
    if status:
        query = query.filter(BOM.status == status)
        
    if customer_name:
        query = query.filter(BOM.customer_name.like(f"%{customer_name}%"))
        
    if prepared_by:
        query = query.filter(BOM.prepared_by.like(f"%{prepared_by}%"))
        
    if created_at_start:
        query = query.filter(BOM.created_at >= created_at_start)
        
    if created_at_end:
        query = query.filter(BOM.created_at <= created_at_end)
    
    # 总数
    total = query.count()
    
    # 分页
    boms = query.order_by(BOM.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    
    return BOMListResponse(
        items=boms,
        total=total,
        page=page,
        page_size=page_size
    )


@router.get("/{bom_id}", response_model=BOMSchema)
async def get_bom(
    bom_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取BOM详情"""
    bom = db.query(BOM).options(joinedload(BOM.items)).filter(BOM.id == bom_id).first()
    if not bom:
        raise HTTPException(status_code=404, detail="BOM不存在")
    return bom


@router.post("", response_model=BOMSchema)
async def create_bom(
    bom_data: BOMCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """创建BOM"""
    try:
        # 如果编码为空，自动生成
        if not bom_data.code or bom_data.code.strip() == "":
            bom_data.code = generate_bom_code(db)
            
        # 检查编码是否已存在
        existing = db.query(BOM).filter(BOM.code == bom_data.code).first()
        if existing:
            raise HTTPException(status_code=400, detail="BOM编码已存在")
        
        # 创建BOM
        bom = BOM(
            code=bom_data.code,
            name=bom_data.name,
            product_name=bom_data.product_name,
            description=bom_data.description,
            status=bom_data.status,
            remark=bom_data.remark,
            customer_name=bom_data.customer_name,
            date=bom_data.date,
            version=bom_data.version,
            sales_channel=bom_data.sales_channel,
            prepared_by=current_user.real_name or current_user.username, # 自动设置为当前用户
            pricing_reviewer=bom_data.pricing_reviewer,
            created_by=current_user.id
        )
        db.add(bom)
        db.flush()
        
        # 创建明细
        total_amount = Decimal("0")
        for item_data in bom_data.items:
            item = BOMItem(
                bom_id=bom.id,
                sequence=item_data.sequence,
                material_name=item_data.material_name,
                specification=item_data.specification,
                unit=item_data.unit,
                quantity=item_data.quantity,
                unit_price=item_data.unit_price,
                total_price=item_data.total_price or (item_data.quantity * item_data.unit_price if item_data.unit_price else None),
                remark=item_data.remark,
                material_category=item_data.material_category,
                material_grade=item_data.material_grade,
                unit_weight=item_data.unit_weight,
                total_weight=item_data.total_weight,
                brand_manufacturer=item_data.brand_manufacturer,
                standard_number=item_data.standard_number,
                surface_treatment=item_data.surface_treatment,
                created_by=bom.created_by  # 使用BOM的创建人
            )
            if item.total_price:
                total_amount += item.total_price
            db.add(item)
        
        bom.total_amount = total_amount
        db.commit()
        db.refresh(bom)
        return bom
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"创建BOM失败: {str(e)}")


@router.put("/{bom_id}", response_model=BOMSchema)
async def update_bom(
    bom_id: int,
    bom_data: BOMUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """更新BOM"""
    bom = db.query(BOM).filter(BOM.id == bom_id).first()
    if not bom:
        raise HTTPException(status_code=404, detail="BOM不存在")
    
    # 更新字段
    if bom_data.name is not None:
        bom.name = bom_data.name
    if bom_data.product_name is not None:
        bom.product_name = bom_data.product_name
    if bom_data.description is not None:
        bom.description = bom_data.description
    if bom_data.status is not None:
        bom.status = bom_data.status
    if bom_data.remark is not None:
        bom.remark = bom_data.remark
    if bom_data.customer_name is not None:
        bom.customer_name = bom_data.customer_name
    if bom_data.date is not None:
        bom.date = bom_data.date
    if bom_data.version is not None:
        bom.version = bom_data.version
    if bom_data.sales_channel is not None:
        bom.sales_channel = bom_data.sales_channel
    if bom_data.prepared_by is not None:
        bom.prepared_by = bom_data.prepared_by
    if bom_data.pricing_reviewer is not None:
        bom.pricing_reviewer = bom_data.pricing_reviewer
    
    # 更新明细
    if bom_data.items is not None:
        # 删除旧明细
        db.query(BOMItem).filter(BOMItem.bom_id == bom_id).delete()
        
        # 添加新明细
        total_amount = Decimal("0")
        for item_data in bom_data.items:
            item = BOMItem(
                bom_id=bom.id,
                sequence=item_data.sequence,
                material_name=item_data.material_name,
                specification=item_data.specification,
                unit=item_data.unit,
                quantity=item_data.quantity,
                unit_price=item_data.unit_price,
                total_price=item_data.total_price or (item_data.quantity * item_data.unit_price if item_data.unit_price else None),
                remark=item_data.remark,
                material_category=item_data.material_category,
                material_grade=item_data.material_grade,
                unit_weight=item_data.unit_weight,
                total_weight=item_data.total_weight,
                brand_manufacturer=item_data.brand_manufacturer,
                standard_number=item_data.standard_number,
                surface_treatment=item_data.surface_treatment,
                created_by=bom.created_by  # 使用BOM的创建人
            )
            if item.total_price:
                total_amount += item.total_price
            db.add(item)
        
        bom.total_amount = total_amount
    
    db.commit()
    db.refresh(bom)
    return bom


@router.delete("/{bom_id}")
async def delete_bom(
    bom_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """删除BOM"""
    bom = db.query(BOM).filter(BOM.id == bom_id).first()
    if not bom:
        raise HTTPException(status_code=404, detail="BOM不存在")
    
    db.delete(bom)
    db.commit()
    return {"message": "删除成功"}


@router.get("/{bom_id}/export")
async def export_bom(
    bom_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """导出BOM为Excel"""
    bom = db.query(BOM).options(joinedload(BOM.items)).filter(BOM.id == bom_id).first()
    if not bom:
        raise HTTPException(status_code=404, detail="BOM不存在")
    
    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM清单"
    
    # 设置标题样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    
    # BOM基本信息
    row = 1
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws.cell(row=row, column=1, value=f"BOM清单 - {bom.name}")
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    row += 2
    ws.cell(row=row, column=1, value="BOM编码:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=bom.code)
    row += 1
    ws.cell(row=row, column=1, value="BOM名称:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=bom.name)
    row += 1
    if bom.product_name:
        ws.cell(row=row, column=1, value="产品名称:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=bom.product_name)
        row += 1
    if bom.description:
        ws.cell(row=row, column=1, value="描述:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=bom.description)
        row += 1
    ws.cell(row=row, column=1, value="状态:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=bom.status or "草稿")
    if bom.customer_name:
        row += 1
        ws.cell(row=row, column=1, value="客户名称:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=bom.customer_name)
    if bom.date:
        row += 1
        ws.cell(row=row, column=1, value="日期:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=bom.date.strftime('%Y-%m-%d') if hasattr(bom.date, 'strftime') else str(bom.date))
    if bom.version:
        row += 1
        ws.cell(row=row, column=1, value="版本号:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=bom.version)
    if bom.sales_channel:
        row += 1
        ws.cell(row=row, column=1, value="销售渠道:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=bom.sales_channel)
    if bom.prepared_by:
        row += 1
        ws.cell(row=row, column=1, value="制单人:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=bom.prepared_by)
    if bom.pricing_reviewer:
        row += 1
        ws.cell(row=row, column=1, value="核价人:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=bom.pricing_reviewer)
    row += 1
    if bom.total_amount:
        ws.cell(row=row, column=1, value="总金额:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=float(bom.total_amount))
        row += 1
    
    # 明细表头
    row += 2
    headers = ["序号", "物料名称", "规格型号", "单位", "数量", "单价", "总价", "物料类别", "材质/牌号", "单重（kg）", "总重（kg）", "品牌/厂家", "标准号/图床", "表面处理", "备注"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 明细数据
    row += 1
    for item in bom.items:
        ws.cell(row=row, column=1, value=item.sequence or "")
        ws.cell(row=row, column=2, value=item.material_name)
        ws.cell(row=row, column=3, value=item.specification or "")
        ws.cell(row=row, column=4, value=item.unit or "")
        ws.cell(row=row, column=5, value=float(item.quantity) if item.quantity else "")
        ws.cell(row=row, column=6, value=float(item.unit_price) if item.unit_price else "")
        ws.cell(row=row, column=7, value=float(item.total_price) if item.total_price else "")
        ws.cell(row=row, column=8, value=item.material_category or "")
        ws.cell(row=row, column=9, value=item.material_grade or "")
        ws.cell(row=row, column=10, value=float(item.unit_weight) if item.unit_weight else "")
        ws.cell(row=row, column=11, value=float(item.total_weight) if item.total_weight else "")
        ws.cell(row=row, column=12, value=item.brand_manufacturer or "")
        ws.cell(row=row, column=13, value=item.standard_number or "")
        ws.cell(row=row, column=14, value=item.surface_treatment or "")
        ws.cell(row=row, column=15, value=item.remark or "")
        row += 1
    
    # 设置列宽
    column_widths = [10, 30, 30, 10, 12, 12, 12, 15, 15, 12, 12, 20, 20, 15, 30]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 简化文件名，避免中文编码问题
    safe_filename = f"BOM_{bom.code}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{safe_filename}"'
        }
    )


@router.post("/import", response_model=ExcelImportResponse)
async def import_bom(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """导入BOM Excel文件"""
    if not file.filename:
        raise HTTPException(status_code=400, detail="文件名不能为空")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400, 
            detail=f"不支持的文件格式。当前文件: {file.filename}，只支持 .xlsx 或 .xls 格式的Excel文件"
        )
    
    # 读取Excel文件
    try:
        contents = await file.read()
        if not contents:
            raise HTTPException(
                status_code=400, 
                detail="文件内容为空。请确保文件已正确上传且文件大小不为0"
            )
        
        # 检查文件大小
        file_size = len(contents)
        if file_size < 100:  # Excel文件最小应该有几KB
            raise HTTPException(
                status_code=400,
                detail=f"文件大小异常（{file_size}字节），可能不是有效的Excel文件。请检查文件是否损坏"
            )
        
        # 检查文件签名（Magic Number）
        # Excel文件的签名：.xlsx文件以PK开头（ZIP格式），.xls文件以特定字节开头
        if file.filename.endswith('.xlsx'):
            if not contents.startswith(b'PK'):
                raise HTTPException(
                    status_code=400,
                    detail="文件格式错误：.xlsx文件应该以ZIP格式开头。可能的原因：1) 文件损坏 2) 文件被重命名但格式不正确 3) 文件不是真正的Excel文件"
                )
        elif file.filename.endswith('.xls'):
            # .xls文件的签名检查
            if contents[:8] != b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
                raise HTTPException(
                    status_code=400,
                    detail="文件格式错误：.xls文件格式不正确。可能的原因：1) 文件损坏 2) 文件被重命名但格式不正确 3) 文件不是真正的Excel文件"
                )
        
        # 尝试加载工作簿
        try:
            if file.filename.endswith('.xls'):
                # .xls格式需要使用xlrd库
                if not HAS_XLRD:
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "系统不支持.xls格式文件。\n\n"
                            "解决方案：\n"
                            "1. 在Excel中打开文件，选择'文件' -> '另存为'\n"
                            "2. 在'文件类型'中选择'Excel工作簿(*.xlsx)'\n"
                            "3. 保存后使用新的.xlsx文件导入\n\n"
                            f"当前文件：{file.filename}\n"
                            "提示：.xls是旧版Excel格式，建议转换为.xlsx格式以获得更好的兼容性"
                        )
                    )
                # 使用xlrd读取.xls文件，然后转换为openpyxl格式
                try:
                    xls_book = xlrd.open_workbook(file_contents=contents)
                    # 创建一个新的openpyxl工作簿
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = xls_book.sheet_names()[0] if xls_book.sheet_names() else "Sheet1"
                    
                    # 读取第一个工作表
                    xls_sheet = xls_book.sheet_by_index(0)
                    
                    # 将数据复制到openpyxl工作表
                    for row_idx in range(xls_sheet.nrows):
                        for col_idx in range(xls_sheet.ncols):
                            cell_value = xls_sheet.cell_value(row_idx, col_idx)
                            # 处理日期类型
                            if xls_sheet.cell_type(row_idx, col_idx) == xlrd.XL_CELL_DATE:
                                try:
                                    date_tuple = xlrd.xldate_as_tuple(cell_value, xls_book.datemode)
                                    cell_value = datetime(*date_tuple[:6])
                                except:
                                    pass  # 如果日期转换失败，保持原值
                            ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
                except Exception as xls_error:
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            f"无法读取.xls格式文件。\n\n"
                            f"错误信息：{str(xls_error)}\n"
                            f"文件：{file.filename}\n\n"
                            "建议：\n"
                            "1. 在Excel中打开文件，确认文件可以正常打开\n"
                            "2. 如果文件可以打开，选择'文件' -> '另存为' -> 保存为.xlsx格式\n"
                            "3. 使用转换后的.xlsx文件导入\n"
                            "4. 检查文件是否损坏"
                        )
                    )
            else:
                # .xlsx格式使用openpyxl直接读取
                wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
        except HTTPException:
            raise
        except Exception as e:
            error_msg = str(e)
            # 提供更详细的错误信息
            if "File contains no valid workbook part" in error_msg:
                detail_msg = (
                    "Excel文件格式错误：文件中没有有效的工作簿部分。\n"
                    "可能的原因：\n"
                    "1. 文件损坏或不完整\n"
                    "2. 文件不是真正的Excel文件（可能是CSV或其他格式被重命名）\n"
                    "3. 文件是用不兼容的软件创建的\n"
                    "4. 文件在传输过程中损坏\n\n"
                    f"技术细节：{error_msg}\n"
                    f"文件大小：{file_size} 字节\n"
                    f"文件名：{file.filename}"
                )
            elif "BadZipFile" in error_msg or "zipfile" in error_msg.lower():
                detail_msg = (
                    "ZIP文件格式错误：.xlsx文件实际上是ZIP压缩包，但ZIP格式不正确。\n"
                    "可能的原因：\n"
                    "1. 文件损坏\n"
                    "2. 文件不是真正的.xlsx文件\n"
                    "3. 文件在传输过程中损坏\n\n"
                    f"技术细节：{error_msg}"
                )
            else:
                detail_msg = (
                    f"无法读取Excel文件。\n"
                    f"错误类型：{type(e).__name__}\n"
                    f"错误信息：{error_msg}\n"
                    f"文件大小：{file_size} 字节\n"
                    f"文件名：{file.filename}\n\n"
                    "建议：\n"
                    "1. 确认文件是使用Microsoft Excel或兼容软件创建的\n"
                    "2. 尝试在Excel中打开文件，确认文件可以正常打开\n"
                    "3. 如果文件可以打开，尝试另存为新文件后再导入"
                )
            raise HTTPException(status_code=400, detail=detail_msg)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=400, 
            detail=f"读取文件时发生未知错误：{str(e)}\n文件：{file.filename if file.filename else '未知'}"
        )
    ws = wb.active
    
    success_count = 0
    error_count = 0
    errors = []
    
    # 查找BOM基本信息（通常在文件开头）
    bom_code = None
    bom_name = None
    product_name = None
    description = None
    status = "草稿"
    customer_name = None
    date = None
    version = None
    prepared_by = None
    pricing_reviewer = None
    sales_channel = None
    bom_remark = None
    
    # 查找明细表头行
    header_row = None
    # 存储列名对应的索引
    col_map = {}
    
    # 将文件内容读取到列表中，以便多次遍历
    all_rows = list(ws.iter_rows(values_only=True))
    
    # 第一遍扫描：查找表头行
    for row_idx, row in enumerate(all_rows):
        row_values = [str(cell).strip() if cell else "" for cell in row]
        row_str = " ".join(row_values).upper()
        
        # 查找表头行（包含"序号"、"设备"、"数量"等关键词）
        # 检查是否包含表头关键词
        keywords = ["序号", "设备", "物料", "名称", "规格", "单位", "数量", "单价", "总价"]
        keyword_count = sum(1 for keyword in keywords if keyword in row_str)
        
        # 至少包含3个关键词才认为是表头（提高准确性）
        if keyword_count >= 3:
            header_row = row_idx
            
            # 构建列映射
            for col_idx, col_name in enumerate(row_values):
                if not col_name:
                    continue
                
                col_name_upper = col_name.upper()
                
                # 序号
                if "序号" in col_name_upper:
                    col_map["sequence"] = col_idx
                # 物料类别（必须在物料名称之前匹配，避免"物料类别"被误匹配为物料名称）
                elif "物料类别" in col_name_upper or "物料分类" in col_name_upper:
                    # 优先匹配"物料类别"或"物料分类"
                    if "material_category" not in col_map:  # 确保还没有匹配过
                        col_map["material_category"] = col_idx
                elif "类别" in col_name_upper and "产品" not in col_name_upper and "物料" not in col_name_upper:
                    # 如果只是"类别"且不包含"产品"和"物料"，也匹配物料类别
                    # 但要排除已经匹配过的"物料类别"
                    if "material_category" not in col_map:  # 确保还没有匹配过
                        col_map["material_category"] = col_idx
                # 物料名称（必须在物料类别之后匹配，避免"物料类别"被误匹配）
                elif any(k in col_name_upper for k in ["设备", "物料", "名称"]) and "客户" not in col_name_upper and "产品" not in col_name_upper and "BOM" not in col_name_upper and "类别" not in col_name_upper and "分类" not in col_name_upper:
                    # 避免匹配到"产品名称"、"客户名称"、"BOM名称"、"物料类别"、"物料分类"
                    if "material_name" not in col_map:  # 确保还没有匹配过
                        col_map["material_name"] = col_idx
                # 规格型号（优先匹配"规格型号"，然后匹配单独的"规格"或"型号"）
                elif "规格型号" in col_name_upper:
                    col_map["specification"] = col_idx
                elif "规格" in col_name_upper or "型号" in col_name_upper:
                    if "specification" not in col_map:  # 确保还没有匹配过
                        col_map["specification"] = col_idx
                # 单位
                elif "单位" in col_name_upper:
                    col_map["unit"] = col_idx
                # 数量
                elif "数量" in col_name_upper:
                    col_map["quantity"] = col_idx
                # 单价
                elif "单价" in col_name_upper:
                    col_map["unit_price"] = col_idx
                # 总价
                elif "总价" in col_name_upper or "总金额" in col_name_upper:
                    col_map["total_price"] = col_idx
                # 材质/牌号（优先匹配"材质/牌号"或"材质牌号"，然后匹配单独的"材质"或"牌号"）
                elif "材质" in col_name_upper and "牌号" in col_name_upper:
                    col_map["material_grade"] = col_idx
                elif "材质" in col_name_upper or "牌号" in col_name_upper:
                    if "material_grade" not in col_map:  # 确保还没有匹配过
                        col_map["material_grade"] = col_idx
                # 单重
                elif "单重" in col_name_upper:
                    col_map["unit_weight"] = col_idx
                # 总重
                elif "总重" in col_name_upper:
                    col_map["total_weight"] = col_idx
                # 品牌/厂家
                elif "品牌" in col_name_upper or "厂家" in col_name_upper:
                    col_map["brand_manufacturer"] = col_idx
                # 标准号/图床
                elif "标准" in col_name_upper or "图号" in col_name_upper or "图床" in col_name_upper:
                    col_map["standard_number"] = col_idx
                # 表面处理（优先匹配"表面处理"，避免误匹配其他包含"处理"的列）
                elif "表面处理" in col_name_upper:
                    col_map["surface_treatment"] = col_idx
                elif "表面" in col_name_upper and "处理" in col_name_upper:
                    # 匹配同时包含"表面"和"处理"的列
                    if "surface_treatment" not in col_map:  # 确保还没有匹配过
                        col_map["surface_treatment"] = col_idx
                # 备注 (需要放在最后匹配，防止匹配到BOM备注)
                elif "备注" in col_name_upper:
                    col_map["remark"] = col_idx
            
            break
    
    # 第二遍扫描：查找基础信息（在表头行之前）
    scan_limit = header_row if header_row is not None else len(all_rows)
    # 定义基础信息字段映射
    basic_info_map = {
        "BOM编码": "bom_code",
        "BOM名称": "bom_name",
        "产品名称": "product_name",
        "项目名称": "product_name",  # 支持"项目名称"作为"产品名称"的别名
        "描述": "description",
        "状态": "status",
        "客户名称": "customer_name",
        "日期": "date",
        "清单日期": "date",  # 支持"清单日期"作为"日期"的别名
        "版本号": "version",
        "制单人": "prepared_by",
        "核价人": "pricing_reviewer",
        "销售渠道": "sales_channel",
        "备注": "bom_remark" # 单独处理BOM级别的备注
    }
    
    for row_idx in range(scan_limit):
        row = all_rows[row_idx]
        row_values = [str(cell).strip() if cell else "" for cell in row]
        
        # 遍历每行，查找键值对
        # 假设格式为：标签 | 值 | 标签 | 值 ...
        for i in range(0, len(row_values) - 1):
            cell_value = row_values[i]
            next_cell_value = row_values[i+1]
            
            if not cell_value:
                continue
                
            # 去除冒号和空格
            label = cell_value.replace(":", "").replace("：", "").strip()
            
            # 检查是否匹配基础信息字段
            for key, field_name in basic_info_map.items():
                if key == label:
                    # 如果匹配，获取下一个单元格的值
                    val = next_cell_value
                    
                    # 特殊处理日期
                    if field_name == "date" and val:
                        try:
                            # 尝试解析日期，这里简单处理，实际可能需要更复杂的解析
                            # 如果是Excel日期对象，openpyxl可能已经转换了，这里得到的str可能是日期格式
                            # 如果是字符串，尝试解析
                            from dateutil import parser
                            parsed_date = parser.parse(val)
                            val = parsed_date
                        except:
                            # 如果解析失败，保留原值或设为None
                            pass
                            
                    # 赋值给对应的变量
                    if field_name == "bom_code": bom_code = val
                    elif field_name == "bom_name": bom_name = val
                    elif field_name == "product_name": product_name = val
                    elif field_name == "description": description = val
                    elif field_name == "status": status = val
                    elif field_name == "customer_name": customer_name = val
                    elif field_name == "date": date = val
                    elif field_name == "version": version = val
                    elif field_name == "prepared_by": prepared_by = val
                    elif field_name == "pricing_reviewer": pricing_reviewer = val
                    elif field_name == "sales_channel": sales_channel = val
                    elif field_name == "bom_remark": bom_remark = val
    
    # 如果没有找到BOM编码或名称，尝试从文件名生成
    if not bom_code:
        # 从文件名提取编码（去除扩展名和特殊字符）
        filename_base = os.path.splitext(file.filename)[0]
        filename_base = os.path.basename(filename_base)  # 去除路径
        # 提取文件名中的编码部分（如果有数字和字母组合）
        code_match = re.search(r'[A-Z0-9]+', filename_base.upper())
        if code_match:
            bom_code = f"BOM_{code_match.group()}"
        else:
            # 使用时间戳生成编码
            bom_code = f"BOM_{datetime.now().strftime('%Y%m%d%H%M%S')}"
    
    if not bom_name:
        # 从文件名提取名称
        filename_base = os.path.splitext(file.filename)[0]
        # 去除路径和扩展名
        bom_name = os.path.basename(filename_base)
        # 清理文件名（去除特殊字符，保留中文、字母、数字）
        bom_name = re.sub(r'[^\w\u4e00-\u9fa5]+', '_', bom_name)
        if not bom_name or bom_name == file.filename or len(bom_name) < 2:
            bom_name = f"BOM清单_{datetime.now().strftime('%Y%m%d')}"
    
    # 检查BOM编码是否已存在
    existing_bom = db.query(BOM).filter(BOM.code == bom_code).first()
    if existing_bom:
        # 如果编码已存在，添加时间戳后缀
        bom_code = f"{bom_code}_{datetime.now().strftime('%H%M%S')}"
    
    # 读取明细数据
    items = []
    if header_row is not None:
        # 确保关键列已映射
        if "material_name" not in col_map:
            # 如果没找到物料名称列，尝试使用默认索引
            # 假设第2列是物料名称
            col_map["material_name"] = 1
            if "sequence" not in col_map: col_map["sequence"] = 0
            if "specification" not in col_map: col_map["specification"] = 2
            if "unit" not in col_map: col_map["unit"] = 3
            if "quantity" not in col_map: col_map["quantity"] = 4
        
        for row_idx in range(header_row + 1, len(all_rows)):
            row = all_rows[row_idx]
            row_values = [cell if cell is not None else "" for cell in row]
            
            # 跳过空行
            if not any(str(c).strip() for c in row_values):
                continue
            
            try:
                # 使用列映射提取数据
                def get_val(key):
                    if key in col_map and col_map[key] < len(row_values):
                        val = row_values[col_map[key]]
                        return str(val).strip() if val is not None else ""
                    return ""
                
                material_name = get_val("material_name")
                
                # 跳过物料名称为空的行
                if not material_name:
                    continue
                
                # 跳过包含基础信息关键词的行（防止把页脚或其他信息当做物料）
                info_labels = ["产品名称", "描述", "状态", "总金额", "BOM编码", "BOM名称", "备注", "客户名称", "合计", "制单", "审核"]
                if material_name in info_labels:
                    continue
                    
                sequence = get_val("sequence")
                specification = get_val("specification")
                unit = get_val("unit")
                
                # 解析数量
                quantity = Decimal("0")
                qty_str = get_val("quantity")
                if qty_str:
                    try:
                        qty_str = qty_str.replace(',', '').replace(' ', '')
                        quantity = Decimal(qty_str)
                    except:
                        pass # 默认为0
                
                # 解析单价
                unit_price = None
                price_str = get_val("unit_price")
                if price_str:
                    try:
                        price_str = price_str.replace(',', '').replace(' ', '').replace('￥', '').replace('$', '').replace('¥', '')
                        unit_price = Decimal(price_str)
                    except:
                        pass
                
                # 解析总价
                total_price = None
                total_str = get_val("total_price")
                if total_str:
                    try:
                        total_str = total_str.replace(',', '').replace(' ', '').replace('￥', '').replace('$', '').replace('¥', '')
                        total_price = Decimal(total_str)
                    except:
                        pass
                
                # 如果没有总价，计算总价
                if not total_price and quantity and unit_price:
                    total_price = quantity * unit_price
                
                # 获取其他扩展属性
                # 辅助函数：将空字符串转换为None
                def to_none_if_empty(val):
                    return val if val else None
                
                remark = to_none_if_empty(get_val("remark"))
                material_category = to_none_if_empty(get_val("material_category"))
                material_grade = to_none_if_empty(get_val("material_grade"))
                
                unit_weight = None
                uw_str = get_val("unit_weight")
                if uw_str:
                    try:
                        unit_weight = Decimal(uw_str.replace(',', '').replace(' ', ''))
                    except: pass
                    
                total_weight = None
                tw_str = get_val("total_weight")
                if tw_str:
                    try:
                        total_weight = Decimal(tw_str.replace(',', '').replace(' ', ''))
                    except: pass
                    
                brand_manufacturer = to_none_if_empty(get_val("brand_manufacturer"))
                standard_number = to_none_if_empty(get_val("standard_number"))
                surface_treatment = to_none_if_empty(get_val("surface_treatment"))
                
                # 处理其他可选字段（统一使用to_none_if_empty处理）
                specification = to_none_if_empty(specification)
                unit = to_none_if_empty(unit)
                sequence = to_none_if_empty(sequence)
                
                items.append({
                    "sequence": sequence,
                    "material_name": material_name,
                    "specification": specification,
                    "unit": unit,
                    "quantity": quantity,
                    "unit_price": unit_price,
                    "total_price": total_price,
                    "remark": remark,
                    "material_category": material_category,
                    "material_grade": material_grade,
                    "unit_weight": unit_weight,
                    "total_weight": total_weight,
                    "brand_manufacturer": brand_manufacturer,
                    "standard_number": standard_number,
                    "surface_treatment": surface_treatment
                })
            except Exception as e:
                errors.append({
                    "row": row_idx + 1,
                    "field": "数据解析",
                    "message": str(e)
                })
                error_count += 1
    
    if not items:
        if not header_row:
            raise HTTPException(
                status_code=400, 
                detail=(
                    "未找到有效的明细数据。\n\n"
                    "可能的原因：\n"
                    "1. 文件中没有找到表头行（应包含'序号'、'物料名称'、'规格型号'等列）\n"
                    "2. 表头行格式不符合要求\n"
                    "3. 数据行全部为空\n\n"
                    f"文件：{file.filename}\n"
                    f"总行数：{ws.max_row}\n\n"
                    "建议：\n"
                    "1. 确认文件包含表头行，表头应包含：序号、物料名称、规格型号、单位、数量等列\n"
                    "2. 确认表头行下方有数据行\n"
                    "3. 检查数据行是否为空"
                )
            )
        else:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"未找到有效的明细数据。\n\n"
                    f"表头行：第{header_row + 1}行\n"
                    f"已扫描行数：{len(all_rows) - header_row - 1}\n\n"
                    "可能的原因：\n"
                    "1. 表头行下方没有数据\n"
                    "2. 数据行的物料名称为空\n"
                    "3. 数据格式不正确\n\n"
                    f"文件：{file.filename}\n\n"
                    "建议：\n"
                    "1. 确认表头行下方有数据\n"
                    "2. 确认数据行的物料名称列不为空\n"
                    "3. 检查数据格式是否正确"
                )
            )
    
    # 创建BOM
    try:
        bom = BOM(
            code=bom_code,
            name=bom_name,
            product_name=product_name,
            description=description,
            status=status,
            customer_name=customer_name,
            date=date,
            version=version,
            prepared_by=prepared_by,
            pricing_reviewer=pricing_reviewer,
            sales_channel=sales_channel,
            remark=bom_remark,
            created_by=current_user.id
        )
        db.add(bom)
        db.flush()
        
        # 创建明细
        total_amount = Decimal("0")
        for item_data in items:
            item = BOMItem(
                bom_id=bom.id,
                sequence=item_data.get("sequence"),
                material_name=item_data["material_name"],
                specification=item_data.get("specification"),
                unit=item_data.get("unit"),
                quantity=item_data["quantity"],
                unit_price=item_data.get("unit_price"),
                total_price=item_data.get("total_price"),
                remark=item_data.get("remark"),
                material_category=item_data.get("material_category"),
                material_grade=item_data.get("material_grade"),
                unit_weight=item_data.get("unit_weight"),
                total_weight=item_data.get("total_weight"),
                brand_manufacturer=item_data.get("brand_manufacturer"),
                standard_number=item_data.get("standard_number"),
                surface_treatment=item_data.get("surface_treatment")
            )
            if item.total_price:
                total_amount += item.total_price
            db.add(item)
        
        bom.total_amount = total_amount
        db.commit()
        db.refresh(bom)  # 刷新对象以获取最新数据
        success_count = 1
        
        # 补全 BOMItem.created_by
        # 重新查询刚创建的BOM（虽然上面refresh了，但items需要更新）
        # 实际上，上面的 items 插入后，如果没有 commit，再 add 可能会有问题，但上面已经 db.flush() 并 add(item) 了
        # 这里因为是导入，bom对象是刚创建的，bom.created_by 是 current_user.id
        
        # 修正：在上面创建 item 时添加 created_by
        # 由于我们已经在循环中创建了 item，我们需要修改上面的循环逻辑
        # 或者在这里执行一个批量更新
        
        db.execute(
            text("UPDATE bom_items SET created_by = :created_by WHERE bom_id = :bom_id"),
            {"created_by": bom.created_by, "bom_id": bom.id}
        )
        db.commit()
        
        # 获取实际值，确保类型正确
        # 在 refresh 后，这些属性应该是实际值，但为了确保类型正确，我们显式获取
        bom_id_value = getattr(bom, 'id', None)
        bom_code_value = getattr(bom, 'code', None)
        bom_name_value = getattr(bom, 'name', None)
        
        # 确保返回的数据类型正确，避免序列化错误
        return ExcelImportResponse(
            success=True,
            bom_id=bom_id_value,
            bom_code=bom_code_value,
            bom_name=bom_name_value,
            items_count=len(items),
            errors=errors,
            error_count=error_count
        )
        
    except HTTPException as e:
        raise
    except Exception as e:
        # 记录详细错误信息，包括堆栈跟踪
        import traceback
        error_trace = traceback.format_exc()
        # 打印完整的错误堆栈到控制台
        print(f"\n{'='*80}")
        print(f"BOM导入错误详情:")
        print(f"错误类型: {type(e).__name__}")
        print(f"错误消息: {str(e)}")
        print(f"文件: {file.filename if file else '未知'}")
        print(f"用户: {current_user.username}")
        print(f"\n完整堆栈跟踪:")
        print(error_trace)
        print(f"{'='*80}\n")
        
        raise HTTPException(
            status_code=500,
            detail=f"导入失败: {str(e)}"
        )


@router.get("/{bom_id}/items/{item_id}/price-history", response_model=BOMItemPriceHistoryResponse)
async def get_item_price_history(
    bom_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取BOM明细项的当前价格"""
    # #region agent log
    _log_debug("bom.py:get_item_price_history", "Function entry", {"bom_id": bom_id, "item_id": item_id}, "A")
    # #endregion
    
    # 验证BOM和明细项存在
    bom = db.query(BOM).filter(BOM.id == bom_id).first()
    if not bom:
        raise HTTPException(status_code=404, detail="BOM不存在")
    
    item = db.query(BOMItem).filter(BOMItem.id == item_id, BOMItem.bom_id == bom_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="BOM明细项不存在")
    
    # #region agent log
    _log_debug("bom.py:get_item_price_history", "BOM item found", {
        "item_id": item.id,
        "material_name": item.material_name,
        "specification": item.specification
    }, "A")
    # #endregion
    
    # 匹配所有物料（可能有多个匹配的记录）
    materials = match_all_materials_by_name_and_spec(
        db,
        item.material_name,
        item.specification
    )
    
    # #region agent log
    _log_debug("bom.py:get_item_price_history", "Materials matched", {
        "match_count": len(materials),
        "material_ids": [m.id for m in materials],
        "material_codes": [m.code for m in materials]
    }, "A")
    # #endregion
    
    material_match = None
    price_history_list = []
    statistics = None
    
    if materials:
        # 物料价格状态映射（用于状态备注）
        status_note_map = {
            'valid': None,  # 有效状态，纳入成本核算，不需要备注
            'pending': '待确认，不纳入成本核算',
            'expired': '已过期，不纳入成本核算',
            'abnormal': '异常，不纳入成本核算',
        }
        
        # 遍历所有匹配的物料，显示每个物料的当前价格
        valid_prices = []  # 用于统计有效价格
        valid_materials = []  # 用于存储有效状态的物料
        
        for material in materials:
            if material.last_price:
                price_status = material.price_status
                status_note = status_note_map.get(price_status) if price_status else None

                # 无效价格状态不展示在历史价格中
                if price_status == "expired":
                    continue
                
                # 构建当前价格记录
                # source字段使用物料的source（供应商名称），如果没有则使用"手动录入"
                material_source = material.source if material.source else "手动录入"
                price_history_list.append(
                    BOMItemPriceHistoryItem(
                        id=material.id,  # 使用物料ID作为标识
                        unit_price=material.last_price,
                        currency=material.currency or "CNY",
                        source=material_source,  # 使用物料的来源（供应商名称）
                        material_type=material.material_type,
                        material_code=material.code,
                        price_status=price_status,
                        status_note=status_note,
                        captured_at=material.updated_at or material.created_at
                    )
                )
                
                # 收集有效价格用于统计
                if price_status == 'valid':
                    valid_prices.append(material.last_price)
                    valid_materials.append(material)
        
        # 只有存在有效价格时，才认为匹配成功
        if valid_materials:
            # 使用第一个有效物料作为主要匹配信息
            first_valid_material = valid_materials[0]
            material_match = MaterialMatchInfo(
                material_id=first_valid_material.id,
                material_code=first_valid_material.code,
                material_name=first_valid_material.name,
                specification=first_valid_material.specification,
                matched=True
            )
        # 如果没有有效价格，material_match 保持为 None（认为未匹配）
        
        # #region agent log
        _log_debug("bom.py:get_item_price_history", "Price history built", {
            "price_count": len(price_history_list),
            "valid_price_count": len(valid_prices)
        }, "A")
        # #endregion
        
        # 计算统计信息
        # 记录数只统计有效状态（valid）的记录，平均值/最高价/最低价基于有效状态的价格
        if valid_prices:
            # 确保所有价格都是 Decimal 类型
            decimal_prices = [Decimal(str(p)) for p in valid_prices]
            statistics = PriceStatistics(
                average=sum(decimal_prices) / len(decimal_prices),
                max=max(decimal_prices),
                min=min(decimal_prices),
                count=len(valid_prices),  # 只统计有效状态的记录数
                latest_price=decimal_prices[0] if decimal_prices else None,  # 第一个是最新的
                latest_date=materials[0].updated_at or materials[0].created_at if materials else None
            )
        else:
            # 如果没有有效价格，不显示统计信息
            statistics = None
    
    # #region agent log
    _log_debug("bom.py:get_item_price_history", "Function exit", {
        "price_history_count": len(price_history_list),
        "has_statistics": statistics is not None
    }, "A")
    # #endregion
    
    return BOMItemPriceHistoryResponse(
        item_id=item.id,
        material_name=item.material_name,
        specification=item.specification,
        material_match=material_match,
        price_history=price_history_list,
        statistics=statistics
    )


@router.post("/{bom_id}/calculate-cost", response_model=BOMCostAnalysisResponse)
async def calculate_bom_cost(
    bom_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """计算BOM的成本价"""
    # 验证BOM存在
    bom = db.query(BOM).options(joinedload(BOM.items)).filter(BOM.id == bom_id).first()
    if not bom:
        raise HTTPException(status_code=404, detail="BOM不存在")
    
    items_cost_info = []
    total_cost = Decimal("0")
    matched_count = 0
    unmatched_count = 0
    total_price_history_count = 0  # 总历史价格记录数
    
    # 统计相关
    unmatched_items = []
    source_stats = {
        "manual": {"count": Decimal("0"), "amount": Decimal("0")},
        "quotation": {"count": Decimal("0"), "amount": Decimal("0")},
        "unknown": {"count": Decimal("0"), "amount": Decimal("0")}
    }
    
    # 遍历所有明细项
    for item in bom.items:
        # 匹配所有物料（可能有多个匹配的记录）
        materials = match_all_materials_by_name_and_spec(
            db,
            item.material_name,
            item.specification
        )
        
        # 收集所有匹配物料的有效价格（price_status == 'valid'）
        valid_prices = []
        valid_materials = []
        all_price_count = 0  # 所有有价格的记录数（包括非有效状态）
        
        for material in materials:
            if material.last_price:
                all_price_count += 1  # 统计所有有价格的记录
                # 只收集有效状态的价格用于成本价计算
                if material.price_status == 'valid':
                    valid_prices.append(material.last_price)
                    valid_materials.append(material)
        
        # 只有存在有效价格时，才认为匹配成功
        matched = len(valid_materials) > 0
        cost_price = None
        cost_total = None
        material_id = None
        price_history_count = 0
        
        if matched:
            matched_count += 1
            # 使用第一个有效物料作为主要匹配信息
            first_valid_material = valid_materials[0]
            
            # #region agent log
            _log_debug("bom.py:calculate_bom_cost", "Matched material details", {
                "item_id": item.id,
                "material_name": item.material_name,
                "valid_count": len(valid_materials),
                "first_type": first_valid_material.material_type,
                "first_source": first_valid_material.source,
                "all_types": [m.material_type for m in valid_materials],
                "all_sources": [m.source for m in valid_materials],
                "all_ids": [m.id for m in valid_materials]
            }, "B")
            # #endregion

            material_id = first_valid_material.id
            price_history_count = len(valid_prices)  # 只统计有效状态的记录数
            
            # 成本价计算：使用所有匹配物料的有效价格的平均值
            # 确保所有价格都是 Decimal 类型
            decimal_prices = [Decimal(str(p)) for p in valid_prices]
            cost_price = sum(decimal_prices) / len(decimal_prices)
            # 计算成本小计
            if item.quantity and cost_price:
                cost_total = item.quantity * cost_price
                total_cost += cost_total
                
                # 统计来源分布 - 按价格记录权重分配
                # 如果一个明细项有N个有效价格记录（来自不同来源），则每个记录贡献 1/N 的计数和金额
                weight = Decimal("1") / Decimal(len(valid_materials))
                
                for m in valid_materials:
                    m_type = m.material_type or "unknown"
                    if m_type not in source_stats:
                        m_type = "unknown"
                    
                    source_stats[m_type]["count"] += weight
                    source_stats[m_type]["amount"] += cost_total * weight
        else:
            unmatched_count += 1
            # 即使没有有效价格，也统计所有有价格的记录数（用于参考）
            price_history_count = all_price_count
        
        # 累计总历史价格记录数
        total_price_history_count += price_history_count
        
        item_cost_info = BOMItemCostInfo(
            item_id=item.id,
            sequence=item.sequence,
            material_name=item.material_name,
            specification=item.specification,
            quantity=item.quantity,
            unit=item.unit,
            matched=matched,
            material_id=material_id,
            cost_price=cost_price,
            cost_total=cost_total,
            price_history_count=price_history_count,  # 该明细项的历史价格记录数
            match_status="已匹配" if matched else "未匹配"
        )
        
        items_cost_info.append(item_cost_info)
        
        if not matched:
            unmatched_items.append(item_cost_info)
    
    # 转换统计结果
    price_source_stats = []
    for source_type, stat in source_stats.items():
        percentage = Decimal("0")
        if total_cost > 0:
            percentage = (stat["amount"] / total_cost) * 100
            
        price_source_stats.append({
            "source_type": source_type,
            "count": stat["count"],
            "total_amount": stat["amount"],
            "percentage": percentage
        })
    
    # 引入新的Schema
    from app.schemas.bom import PriceSourceStat
    
    return BOMCostAnalysisResponse(
        bom_id=bom.id,
        bom_code=bom.code,
        bom_name=bom.name,
        total_cost=total_cost,
        items_count=len(bom.items),
        matched_count=matched_count,
        unmatched_count=unmatched_count,
        items=items_cost_info,
        total_price_history_count=total_price_history_count,  # 所有明细项的历史价格记录总数
        price_source_stats=[PriceSourceStat(**s) for s in price_source_stats],
        unmatched_items=unmatched_items
    )


@router.get("/{bom_id}/compare-quotations", response_model=QuotationComparisonResponse)
async def compare_quotations(
    bom_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """对比BOM关联的所有报价单"""
    # 验证BOM存在并加载明细
    bom = db.query(BOM).options(joinedload(BOM.items)).filter(BOM.id == bom_id).first()
    if not bom:
        raise HTTPException(status_code=404, detail="BOM不存在")
    
    # 获取该BOM关联的所有报价单（排除已拒绝的）
    quotations = db.query(Quotation).options(
        joinedload(Quotation.supplier),
        joinedload(Quotation.items)
    ).filter(
        Quotation.bom_id == bom_id,
        Quotation.status != "rejected"  # 排除已拒绝的报价单
    ).order_by(Quotation.created_at.desc()).all()
    
    if not quotations:
        raise HTTPException(status_code=400, detail="该BOM没有可对比的报价单（已拒绝的报价单不纳入对比）")
    
    # 构建报价单基本信息
    quotation_basic_infos = []
    for q in quotations:
        quotation_basic_infos.append(QuotationBasicInfo(
            id=q.id,
            code=q.code,
            supplier_id=q.supplier_id,
            supplier_name=q.supplier.name if q.supplier else "",
            supplier_code=q.supplier.code if q.supplier else None,
            credit_rating=q.supplier.credit_rating if q.supplier else None,
            total_amount=q.total_amount,
            status=q.status,
            quotation_date=q.quotation_date,
            valid_until=q.valid_until,
            delivery_days=q.delivery_days,
            delivery_terms=q.delivery_terms,
            payment_terms=q.payment_terms,
            currency=q.currency
        ))
    
    # 构建BOM物料索引（用于匹配）
    bom_item_index_by_seq = {}  # 按序号索引
    bom_item_index_by_name_spec = {}  # 按名称+规格索引
    
    for bom_item in bom.items:
        if bom_item.sequence:
            bom_item_index_by_seq[bom_item.sequence] = bom_item
        
        # 构建名称+规格索引（去除空格）
        name_clean = _remove_all_spaces(bom_item.material_name)
        spec_clean = _remove_all_spaces(bom_item.specification) if bom_item.specification else ""
        key = f"{name_clean}|{spec_clean}"
        if key not in bom_item_index_by_name_spec:
            bom_item_index_by_name_spec[key] = bom_item
    
    # 为每个BOM物料匹配报价单物料
    item_rows = []
    unmatched_quotations = {q.id: [] for q in quotations}
    
    for bom_item in bom.items:
        cells = []
        
        # 对每个报价单，查找匹配的物料
        for quotation in quotations:
            matched_item = None
            
            # 优先按序号匹配
            if bom_item.sequence:
                for q_item in quotation.items:
                    if q_item.sequence == bom_item.sequence:
                        matched_item = q_item
                        break
            
            # 如果序号未匹配，按名称+规格匹配
            if not matched_item:
                bom_name_clean = _remove_all_spaces(bom_item.material_name)
                bom_spec_clean = _remove_all_spaces(bom_item.specification) if bom_item.specification else ""
                
                for q_item in quotation.items:
                    q_name_clean = _remove_all_spaces(q_item.material_name)
                    q_spec_clean = _remove_all_spaces(q_item.specification) if q_item.specification else ""
                    
                    if bom_name_clean == q_name_clean and bom_spec_clean == q_spec_clean:
                        matched_item = q_item
                        break
            
            # 构建对比单元格
            if matched_item:
                cells.append(ComparisonItemCell(
                    quotation_id=quotation.id,
                    quotation_code=quotation.code,
                    supplier_name=quotation.supplier.name if quotation.supplier else "",
                    unit_price=matched_item.unit_price,
                    total_price=matched_item.total_price,
                    quantity=matched_item.quantity,
                    brand=matched_item.brand,
                    delivery_days=matched_item.delivery_days,
                    remark=matched_item.remark,
                    matched=True
                ))
            else:
                # 未匹配
                cells.append(ComparisonItemCell(
                    quotation_id=quotation.id,
                    quotation_code=quotation.code,
                    supplier_name=quotation.supplier.name if quotation.supplier else "",
                    matched=False
                ))
        
        # 检查该BOM物料是否在所有报价单中都未匹配
        all_unmatched = all(not cell.matched for cell in cells)
        if all_unmatched:
            error_msg = f"BOM物料在所有报价单中都未匹配：序号 {bom_item.sequence or '无'}, 物料名称 {bom_item.material_name}"
            if bom_item.specification:
                error_msg += f", 规格 {bom_item.specification}"
            raise HTTPException(status_code=400, detail=error_msg)
        
        item_rows.append(ComparisonItemRow(
            bom_item_id=bom_item.id,
            sequence=bom_item.sequence,
            material_name=bom_item.material_name,
            specification=bom_item.specification,
            unit=bom_item.unit,
            bom_quantity=bom_item.quantity,
            cells=cells
        ))
    
    # 检查是否有未匹配的报价单物料（报价单中有但BOM中没有的）
    unmatched_items_list = []
    for quotation in quotations:
        for q_item in quotation.items:
            matched = False
            
            # 检查是否匹配到BOM物料
            for bom_item in bom.items:
                # 按序号匹配
                if bom_item.sequence and q_item.sequence == bom_item.sequence:
                    matched = True
                    break
                
                # 按名称+规格匹配
                bom_name_clean = _remove_all_spaces(bom_item.material_name)
                bom_spec_clean = _remove_all_spaces(bom_item.specification) if bom_item.specification else ""
                q_name_clean = _remove_all_spaces(q_item.material_name)
                q_spec_clean = _remove_all_spaces(q_item.specification) if q_item.specification else ""
                
                if bom_name_clean == q_name_clean and bom_spec_clean == q_spec_clean:
                    matched = True
                    break
            
            if not matched:
                unmatched_items_list.append({
                    "quotation_code": quotation.code,
                    "supplier_name": quotation.supplier.name if quotation.supplier else "",
                    "sequence": q_item.sequence,
                    "material_name": q_item.material_name,
                    "specification": q_item.specification,
                })
    
    # 如果有未匹配的物料，报错
    if unmatched_items_list:
        error_msg = "存在未匹配的物料，无法进行对比：\n"
        for item in unmatched_items_list:
            error_msg += f"报价单 {item['quotation_code']} ({item['supplier_name']}): 序号 {item['sequence']}, 物料 {item['material_name']}"
            if item['specification']:
                error_msg += f", 规格 {item['specification']}"
            error_msg += "\n"
        raise HTTPException(status_code=400, detail=error_msg.strip())
    
    # 计算最优标记
    best_markers = {
        "lowest_total_price": None,  # 最低总价报价单ID
        "shortest_delivery_days": None,  # 最短交期报价单ID
        "longest_valid_until": None,  # 最长有效期报价单ID
        "item_lowest_price": {}  # 每个BOM物料的最低单价报价单ID {bom_item_id: quotation_id}
    }
    
    # 找出最低总价
    min_total = None
    min_total_quotation_id = None
    for q in quotations:
        if q.total_amount:
            if min_total is None or q.total_amount < min_total:
                min_total = q.total_amount
                min_total_quotation_id = q.id
    best_markers["lowest_total_price"] = min_total_quotation_id
    
    # 找出最短交期
    min_delivery = None
    min_delivery_quotation_id = None
    for q in quotations:
        if q.delivery_days is not None:
            if min_delivery is None or q.delivery_days < min_delivery:
                min_delivery = q.delivery_days
                min_delivery_quotation_id = q.id
    best_markers["shortest_delivery_days"] = min_delivery_quotation_id
    
    # 找出最长有效期
    max_valid_until = None
    max_valid_quotation_id = None
    for q in quotations:
        if q.valid_until:
            if max_valid_until is None or q.valid_until > max_valid_until:
                max_valid_until = q.valid_until
                max_valid_quotation_id = q.id
    best_markers["longest_valid_until"] = max_valid_quotation_id
    
    # 找出每个物料的最低单价
    for row in item_rows:
        min_price = None
        min_price_quotation_id = None
        for cell in row.cells:
            if cell.matched and cell.unit_price is not None:
                if min_price is None or cell.unit_price < min_price:
                    min_price = cell.unit_price
                    min_price_quotation_id = cell.quotation_id
        if min_price_quotation_id:
            best_markers["item_lowest_price"][row.bom_item_id] = min_price_quotation_id
    
    return QuotationComparisonResponse(
        bom_id=bom.id,
        bom_code=bom.code,
        bom_name=bom.name,
        quotations=quotation_basic_infos,
        item_rows=item_rows,
        unmatched_quotations={},  # 不再返回未匹配物料，因为会直接报错
        best_markers=best_markers
    )


@router.post("/{bom_id}/compare-quotations/export")
async def export_quotation_comparison(
    bom_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """导出报价单对比结果到Excel"""
    # 先获取对比数据
    comparison_data = await compare_quotations(bom_id, db, current_user)
    
    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "报价对比"
    
    # 定义样式
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    best_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 绿色背景
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # 灰色背景
    center_alignment = Alignment(horizontal="center", vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")
    
    row = 1
    
    # 第一行：BOM基本信息
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = f"BOM编码：{comparison_data.bom_code} | BOM名称：{comparison_data.bom_name}"
    ws[f'A{row}'].font = title_font
    row += 1
    
    # 第二行：报价单基本信息对比表头
    ws[f'A{row}'] = "项目"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = center_alignment
    
    col = 2
    for quotation in comparison_data.quotations:
        col_letter = get_column_letter(col)
        ws[f'{col_letter}{row}'] = f"{quotation.supplier_name}\n({quotation.code})"
        ws[f'{col_letter}{row}'].font = header_font
        ws[f'{col_letter}{row}'].fill = header_fill
        ws[f'{col_letter}{row}'].alignment = center_alignment
        ws.column_dimensions[col_letter].width = 20
        col += 1
    
    row += 1
    
    # 报价单基本信息行
    info_rows = [
        ("信用等级", [q.credit_rating or "-" for q in comparison_data.quotations], None),
        ("总价", [float(q.total_amount) if q.total_amount else None for q in comparison_data.quotations], "lowest_total_price"),
        ("交货天数", [q.delivery_days for q in comparison_data.quotations], "shortest_delivery_days"),
        ("有效期至", [q.valid_until.strftime("%Y-%m-%d") if q.valid_until else None for q in comparison_data.quotations], "longest_valid_until"),
        ("付款条件", [q.payment_terms for q in comparison_data.quotations], None),
        ("交货条件", [q.delivery_terms for q in comparison_data.quotations], None),
        ("状态", [q.status for q in comparison_data.quotations], None),
    ]
    
    for info_label, info_values, marker_key in info_rows:
        ws[f'A{row}'] = info_label
        ws[f'A{row}'].font = Font(bold=True)
        col = 2
        for idx, value in enumerate(info_values):
            col_letter = get_column_letter(col)
            ws[f'{col_letter}{row}'] = value if value is not None else "-"
            
            # 标记最优项
            if marker_key and comparison_data.best_markers.get(marker_key) == comparison_data.quotations[idx].id:
                ws[f'{col_letter}{row}'].fill = best_fill
            
            if isinstance(value, (int, float)):
                ws[f'{col_letter}{row}'].alignment = right_alignment
            else:
                ws[f'{col_letter}{row}'].alignment = center_alignment
            
            col += 1
        row += 1
    
    row += 1
    
    # 明细项对比表头
    ws[f'A{row}'] = "序号"
    ws[f'B{row}'] = "物料名称"
    ws[f'C{row}'] = "规格型号"
    ws[f'D{row}'] = "单位"
    ws[f'E{row}'] = "数量"
    
    col = 6
    for quotation in comparison_data.quotations:
        col_letter = get_column_letter(col)
        ws[f'{col_letter}{row}'] = f"{quotation.supplier_name}\n单价"
        ws[f'{get_column_letter(col+1)}{row}'] = f"{quotation.supplier_name}\n总价"
        ws[f'{get_column_letter(col+2)}{row}'] = f"{quotation.supplier_name}\n品牌"
        ws[f'{get_column_letter(col+3)}{row}'] = f"{quotation.supplier_name}\n交期"
        
        for offset in range(4):
            cell = ws[f'{get_column_letter(col+offset)}{row}']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        col += 4
    
    # 设置固定列宽度
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    
    for i in range(6, col):
        ws.column_dimensions[get_column_letter(i)].width = 15
    
    row += 1
    
    # 明细项数据
    for item_row in comparison_data.item_rows:
        ws[f'A{row}'] = item_row.sequence or ""
        ws[f'B{row}'] = item_row.material_name
        ws[f'C{row}'] = item_row.specification or ""
        ws[f'D{row}'] = item_row.unit or ""
        ws[f'E{row}'] = float(item_row.bom_quantity) if item_row.bom_quantity else 0
        ws[f'E{row}'].alignment = right_alignment
        
        col = 6
        for cell in item_row.cells:
            col_letter = get_column_letter(col)
            
            # 单价
            if cell.matched and cell.unit_price is not None:
                ws[f'{col_letter}{row}'] = float(cell.unit_price)
                ws[f'{col_letter}{row}'].number_format = '#,##0.00'
                
                # 标记最低价
                if comparison_data.best_markers.get("item_lowest_price", {}).get(item_row.bom_item_id) == cell.quotation_id:
                    ws[f'{col_letter}{row}'].fill = best_fill
            else:
                ws[f'{col_letter}{row}'] = "-"
            ws[f'{col_letter}{row}'].alignment = right_alignment
            col += 1
            
            # 总价
            if cell.matched and cell.total_price is not None:
                ws[f'{get_column_letter(col)}{row}'] = float(cell.total_price)
                ws[f'{get_column_letter(col)}{row}'].number_format = '#,##0.00'
            else:
                ws[f'{get_column_letter(col)}{row}'] = "-"
            ws[f'{get_column_letter(col)}{row}'].alignment = right_alignment
            col += 1
            
            # 品牌
            ws[f'{get_column_letter(col)}{row}'] = cell.brand or "-"
            ws[f'{get_column_letter(col)}{row}'].alignment = center_alignment
            col += 1
            
            # 交期
            if cell.matched and cell.delivery_days is not None:
                ws[f'{get_column_letter(col)}{row}'] = cell.delivery_days
            else:
                ws[f'{get_column_letter(col)}{row}'] = "-"
            ws[f'{get_column_letter(col)}{row}'].alignment = right_alignment
            col += 1
        
        row += 1
    
    # 冻结首行和左侧列
    ws.freeze_panes = 'F2'
    
    # 生成Excel文件
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 生成文件名
    filename = f"报价对比_{comparison_data.bom_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # 使用RFC 5987标准编码文件名，支持中文
    # 同时提供ASCII兼容的文件名作为fallback，确保更好的兼容性
    safe_filename = f"QuotationComparison_{comparison_data.bom_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = quote(filename.encode('utf-8'))
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{safe_filename}"; filename*=UTF-8\'\'{encoded_filename}'
        }
    )
